import os
import json
from datetime import datetime
from functools import wraps
from flask import (
    Flask, render_template, request, jsonify,
    send_file, send_from_directory, redirect,
    url_for, session, flash
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.config['SECRET_KEY'] = 'animals-samastipur-astrology-care-key-2026'
app.jinja_env.globals.update(enumerate=enumerate)

# --- Credentials ---
ADMIN_USERNAME = 'admin'
ADMIN_PASSWORD = 'Spj@0812'

# --- File Paths ---
DATA_DIR = os.path.join(app.root_path, 'data')
UPDATES_FILE = os.path.join(DATA_DIR, 'updates.json')
EXCEL_FILE = os.path.join(DATA_DIR, 'donations.xlsx')
VOLUNTEER_EXCEL = os.path.join(DATA_DIR, 'volunteers.xlsx')

os.makedirs(DATA_DIR, exist_ok=True)


# ──────────────────────────────────────────────────────────────
# Excel initializers
# ──────────────────────────────────────────────────────────────

def init_donations_excel():
    """Ensure astrology queries excel file exists with formatted headers."""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Astrology Queries"

        headers = [
            "Submission ID", "Date & Time", "Full Name",
            "Date of Birth", "Time of Birth", "Place of Birth",
            "Astrology Question", "Email Address", "Report Status"
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color="1E1E2F", end_color="1E1E2F", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFD700")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = thin_border

        col_widths = [15, 20, 22, 15, 15, 20, 38, 25, 18]
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

        wb.save(EXCEL_FILE)


def init_volunteers_excel():
    """Ensure volunteers excel file exists with headers."""
    if not os.path.exists(VOLUNTEER_EXCEL):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Volunteers"
        headers = ["Submission Date", "Full Name", "Phone", "Email", "Area of Interest", "Message"]
        ws.append(headers)

        header_fill = PatternFill(start_color="1E1E2F", end_color="1E1E2F", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFD700")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = thin_border

        col_widths = [20, 22, 18, 28, 22, 35]
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

        wb.save(VOLUNTEER_EXCEL)


init_donations_excel()
init_volunteers_excel()


# ──────────────────────────────────────────────────────────────
# Auth helpers
# ──────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated


# ──────────────────────────────────────────────────────────────
# Public routes
# ──────────────────────────────────────────────────────────────

@app.route('/img/<path:filename>')
def serve_img(filename):
    return send_from_directory(os.path.join(app.root_path, 'img'), filename)


@app.route('/')
def index():
    updates = []
    if os.path.exists(UPDATES_FILE):
        try:
            with open(UPDATES_FILE, 'r', encoding='utf-8') as f:
                updates = json.load(f)
        except Exception:
            updates = []
    return render_template('index.html', updates=updates)


@app.route('/api/updates', methods=['GET'])
def get_updates():
    if os.path.exists(UPDATES_FILE):
        with open(UPDATES_FILE, 'r', encoding='utf-8') as f:
            updates = json.load(f)
        return jsonify({"success": True, "updates": updates})
    return jsonify({"success": True, "updates": []})


@app.route('/api/submit-astrology', methods=['POST'])
@app.route('/api/donate-astrology', methods=['POST'])
def submit_astrology():
    try:
        data = request.json or request.form

        name = data.get('name', '').strip()
        dob = data.get('dob', '').strip()
        tob = data.get('tob', '').strip()
        pob = data.get('pob', '').strip()
        question = data.get('question', '').strip()
        email = data.get('email', '').strip()

        if not name or not email or not question:
            return jsonify({"success": False, "message": "Please fill all required fields (Name, Email, Question)."}), 400

        init_donations_excel()
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active

        sub_id = f"ASTRO-{int(datetime.now().timestamp())}"
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        row_data = [sub_id, now_str, name, dob, tob, pob, question, email, "Pending PDF Report"]
        ws.append(row_data)

        row_idx = ws.max_row
        thin_border = Border(
            left=Side(style='thin', color='EEEEEE'),
            right=Side(style='thin', color='EEEEEE'),
            top=Side(style='thin', color='EEEEEE'),
            bottom=Side(style='thin', color='EEEEEE')
        )
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx in [1, 2, 4, 5, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        wb.save(EXCEL_FILE)

        return jsonify({
            "success": True,
            "message": f"Thank you, {name}! Your astrological query has been submitted successfully. Our team will prepare your Kundali PDF report and email it to {email} along with details on how you can support Animals of Samastipur.",
            "submission_id": sub_id
        })
    except Exception as e:
        return jsonify({"success": False, "message": f"Error saving details: {str(e)}"}), 500


@app.route('/api/volunteer', methods=['POST'])
def register_volunteer():
    try:
        data = request.json or request.form
        name = data.get('name', '').strip()
        phone = data.get('phone', '').strip()
        email = data.get('email', '').strip()
        interest = data.get('interest', '').strip()
        message = data.get('message', '').strip()

        if not name or not phone or not email:
            return jsonify({"success": False, "message": "Name, Phone, and Email are required."}), 400

        init_volunteers_excel()
        wb = openpyxl.load_workbook(VOLUNTEER_EXCEL)
        ws = wb.active
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([now_str, name, phone, email, interest, message])
        wb.save(VOLUNTEER_EXCEL)

        return jsonify({
            "success": True,
            "message": f"Welcome to the Animals of Samastipur family, {name}! We will reach out to you shortly on {phone}."
        })
    except Exception as e:
        return jsonify({"success": False, "message": f"Error registering volunteer: {str(e)}"}), 500


# ──────────────────────────────────────────────────────────────
# Admin – Login / Logout
# ──────────────────────────────────────────────────────────────

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if session.get('admin_logged_in'):
        return redirect(url_for('admin_dashboard'))

    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            session.permanent = False
            return redirect(url_for('admin_dashboard'))
        else:
            error = 'Invalid username or password. Please try again.'

    return render_template('login.html', error=error)


@app.route('/admin/logout')
def admin_logout():
    session.clear()
    return redirect(url_for('admin_login'))


# ──────────────────────────────────────────────────────────────
# Admin – Dashboard (both tables)
# ──────────────────────────────────────────────────────────────

@app.route('/admin')
@app.route('/admin/dashboard')
@app.route('/admin/donations')
@app.route('/admin/volunteers')
@login_required
def admin_dashboard():
    # Donations / Astrology queries
    init_donations_excel()
    wb_d = openpyxl.load_workbook(EXCEL_FILE)
    ws_d = wb_d.active
    d_rows = list(ws_d.iter_rows(values_only=True))
    donation_headers = list(d_rows[0]) if d_rows else []
    donation_rows = [list(r) for r in d_rows[1:]] if len(d_rows) > 1 else []

    # Volunteers
    init_volunteers_excel()
    wb_v = openpyxl.load_workbook(VOLUNTEER_EXCEL)
    ws_v = wb_v.active
    v_rows = list(ws_v.iter_rows(values_only=True))
    volunteer_headers = list(v_rows[0]) if v_rows else []
    volunteer_rows = [list(r) for r in v_rows[1:]] if len(v_rows) > 1 else []

    return render_template(
        'admin.html',
        donation_headers=donation_headers,
        donation_rows=donation_rows,
        volunteer_headers=volunteer_headers,
        volunteer_rows=volunteer_rows,
        total_donations=len(donation_rows),
        total_volunteers=len(volunteer_rows),
    )


# ──────────────────────────────────────────────────────────────
# Admin – Downloads
# ──────────────────────────────────────────────────────────────

@app.route('/download/donations')
@login_required
def download_donations():
    init_donations_excel()
    return send_file(
        EXCEL_FILE,
        as_attachment=True,
        download_name="Animals_of_Samastipur_Astrology_Queries.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route('/download/volunteers')
@login_required
def download_volunteers():
    init_volunteers_excel()
    return send_file(
        VOLUNTEER_EXCEL,
        as_attachment=True,
        download_name="Animals_of_Samastipur_Volunteers.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == '__main__':
    print("Starting Animals of Samastipur Web App on http://127.0.0.1:5000")
    app.run(debug=True, port=5000)
